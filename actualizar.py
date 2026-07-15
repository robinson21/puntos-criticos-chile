#!/usr/bin/env python3
"""
Actualizador de datos Puntos Críticos SENAPRED.
Descarga el Excel, procesa, genera JSONs regionales y hace deploy a gh-pages.

Uso:
  python3 actualizar.py          # Solo descarga y procesa
  python3 actualizar.py --push   # Descarga, procesa y hace deploy
"""

import json, os, gzip, sys, subprocess
from collections import defaultdict
from urllib.request import urlopen
from urllib.error import URLError

PROJECT_DIR = '/home/ubuntu/puntos-criticos'
PUBLIC_DIR = os.path.join(PROJECT_DIR, 'public', 'data')
SRC_DATA = os.path.join(PROJECT_DIR, 'src', 'data')

# URLs oficiales donde SENAPRED publica el Excel
# Verificar cada año cuál es la URL vigente
URLS = [
    'https://web.senapred.cl/wp-content/uploads/2026/04/Puntos_Criticos_INV_2026.xlsx',
    'https://senapred.cl/wp-content/uploads/2026/04/Puntos_Criticos_INV_2026.xlsx',
]

# Fallback: el repositorio original
FALLBACK_URL = 'https://raw.githubusercontent.com/psarricolea/puntos-criticos-chile/main/Puntos_Criticos_INV_2026.xlsx'

REGION_SLUGS = {
    'METROPOLITANA DE SANTIAGO': 'metropolitana_de_santiago',
    'VALPARAÍSO': 'valparaiso',
    'BIOBÍO': 'biobio',
    'COQUIMBO': 'coquimbo',
    'LA ARAUCANÍA': 'la_araucania',
    'LIBERTADOR GENERAL BERNARDO OHIGGINS': 'libertador_general_bernardo_ohiggins',
    'MAULE': 'maule',
    'LOS LAGOS': 'los_lagos',
    'ÑUBLE': 'nuble',
    'ATACAMA': 'atacama',
    'LOS RÍOS': 'los_rios',
    'ANTOFAGASTA': 'antofagasta',
    'AYSÉN DEL GENERAL CARLOS IBANEZ DEL CAMPO': 'aysen_del_general_carlos_ibanez_del_campo',
    'MAGALLANES Y DE LA ANTÁRTICA CHILENA': 'magallanes_y_de_la_antartica_chilena',
}


def descargar_excel():
    """Descarga el Excel de SENAPRED. Prueba URLs oficiales, luego fallback."""
    for url in URLS + [FALLBACK_URL]:
        try:
            print(f'  Intentando: {url}')
            resp = urlopen(url, timeout=30)
            data = resp.read()
            if len(data) > 10000:
                print(f'  ✅ Descargado: {len(data)/1024:.0f} KB')
                return data
        except Exception as e:
            print(f'  ❌ {e}')
    raise Exception('No se pudo descargar el Excel desde ninguna URL')


def procesar_excel(excel_data):
    """Procesa el Excel y genera los JSONs regionales + metadata."""
    import openpyxl
    from io import BytesIO

    wb = openpyxl.load_workbook(BytesIO(excel_data), read_only=True, data_only=True)
    sheet = wb.active if wb.active else wb[wb.sheetnames[0]]

    records = []
    total_rows = 0
    for row in sheet.iter_rows(min_row=2, values_only=True):
        total_rows += 1
        lon, lat = row[13], row[14]
        if lon is None or lat is None:
            continue
        region = str(row[0] or '').strip().upper()
        records.append({
            'c': str(row[2] or '').strip().upper(),   # comuna
            'cs': str(row[4] or '').strip(),            # causa
            'lv': str(row[10] or '').strip(),           # nivel riesgo
            'tp': str(row[11] or '').strip(),           # tipo solucion
            'lat': round(float(lat), 4),
            'lon': round(float(lon), 4),
        })
        # Asignar region después
        records[-1]['r'] = region

    print(f'  Filas leídas: {total_rows}')
    print(f'  Registros con coordenadas: {len(records)}')

    # Agrupar por región
    by_region = defaultdict(list)
    for r in records:
        by_region[r['r']].append(r)

    # Crear directorios
    os.makedirs(PUBLIC_DIR, exist_ok=True)

    # Escribir archivos regionales
    region_map = {}
    for region, items in sorted(by_region.items(), key=lambda x: len(x[1]), reverse=True):
        slug = REGION_SLUGS.get(region, region.lower().replace(' ','_'))
        clean = [{'c':i['c'],'cs':i['cs'],'lv':i['lv'],'tp':i['tp'],'lat':i['lat'],'lon':i['lon']} for i in items]
        path = os.path.join(PUBLIC_DIR, f'{slug}.json')
        with open(path, 'w') as f:
            json.dump(clean, f, ensure_ascii=False)
        region_map[region] = slug
        print(f'  {slug}: {len(clean)} pts')

    # Metadata global
    comuna_stats = defaultdict(lambda: {'n': 0})
    for r in records:
        comuna_stats[r['c']]['n'] += 1
        if 'region' not in comuna_stats[r['c']]:
            comuna_stats[r['c']]['region'] = r['r']

    ranking = sorted(
        [{'c': c, 'n': v['n'], 'r': v['region'], 'idx': min(100, round(v['n'] / 4.21))}
         for c, v in comuna_stats.items()],
        key=lambda x: x['idx'], reverse=True
    )

    niveles = defaultdict(int)
    causas = defaultdict(int)
    tipos = defaultdict(int)
    for r in records:
        niveles[r['lv']] += 1
        causas[r['cs']] += 1
        tipos[r['tp']] += 1

    metadata = {
        'total': len(records),
        'regiones': region_map,
        'ranking': ranking[:100],
        'niveles': dict(niveles),
        'causas': dict(causas),
        'tipos': dict(tipos),
        'comunas': len(comuna_stats),
    }

    with open(os.path.join(PUBLIC_DIR, 'meta.json'), 'w') as f:
        json.dump(metadata, f, ensure_ascii=False)

    return metadata


def hacer_deploy():
    """Hace build y push a gh-pages."""
    os.chdir(PROJECT_DIR)
    result = subprocess.run(['npm', 'run', 'build'], capture_output=True, text=True, timeout=120)
    if result.returncode != 0:
        print(f'Build failed: {result.stderr}')
        return False
    result = subprocess.run(['npx', 'gh-pages', '-d', 'dist'], capture_output=True, text=True, timeout=120)
    if result.returncode == 0:
        print('✅ Deploy exitoso')
        return True
    else:
        print(f'Deploy falló: {result.stderr}')
        return False


if __name__ == '__main__':
    do_push = '--push' in sys.argv

    print('=== Actualizador Puntos Críticos SENAPRED ===')
    print('\n1️⃣  Descargando Excel...')
    excel = descargar_excel()

    print('\n2️⃣  Procesando datos...')
    meta = procesar_excel(excel)
    print(f'   ✅ {meta["total"]} registros · {meta["comunas"]} comunas · {len(meta["regiones"])} regiones')

    if do_push:
        print('\n3️⃣  Build y deploy...')
        hacer_deploy()
    else:
        print('\n3️⃣  Skipeando deploy (usa --push para hacerlo)')
        print('   Datos actualizados en public/data/')

    print('\n✅ Listo')
